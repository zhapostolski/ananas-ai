#!/usr/bin/env python3
"""
context/ananas/raw/ watcher
Detects new files dropped into the raw folder, extracts content,
and updates ananas-overview.md via Claude API automatically.

Usage:
    python scripts/watch_context.py          # start watching
    python scripts/watch_context.py --once   # process pending files and exit
"""

import argparse
import base64
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

# ── paths ──────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "context" / "ananas" / "raw"
OVERVIEW = ROOT / "context" / "ananas" / "ananas-overview.md"
SOURCES = ROOT / "context" / "ananas" / "SOURCES.md"
PROCESSED_DIR = RAW_DIR / ".processed"

# ── logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("context-watcher")

# ── supported file types ───────────────────────────────────────────────────────
TEXT_SUFFIXES = {".txt", ".md", ".json", ".csv", ".html", ".xml"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
MEDIA_TYPE = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}


# ── file readers ───────────────────────────────────────────────────────────────

def read_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = [p.extract_text() or "" for p in reader.pages]
    return "\n\n".join(pages)


def read_docx(path: Path) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        rows.append(f"## Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            if any(c is not None for c in row):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(rows)


def extract_content(path: Path) -> dict:
    """
    Returns a dict with:
      - type: "text" | "image"
      - content: str (for text) or base64 str (for image)
      - media_type: str (for image only)
    """
    suffix = path.suffix.lower()

    if suffix in TEXT_SUFFIXES:
        return {"type": "text", "content": path.read_text(encoding="utf-8", errors="replace")}

    if suffix == ".pdf":
        return {"type": "text", "content": read_pdf(path)}

    if suffix in {".docx"}:
        return {"type": "text", "content": read_docx(path)}

    if suffix in {".xlsx", ".xls"}:
        return {"type": "text", "content": read_xlsx(path)}

    if suffix in IMAGE_SUFFIXES:
        data = base64.standard_b64encode(path.read_bytes()).decode()
        return {"type": "image", "content": data, "media_type": MEDIA_TYPE[suffix]}

    # fallback: try reading as text
    try:
        return {"type": "text", "content": path.read_text(encoding="utf-8", errors="replace")}
    except Exception:
        return {"type": "text", "content": f"[Could not read file: {path.name}]"}


# ── Claude API call ────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are maintaining a living company overview document for Ananas, an e-commerce company.
You will receive:
1. The current state of ananas-overview.md
2. One or more new source documents/images dropped by the team

Your job is to extract all relevant company information from the source materials and update the overview.

Rules:
- Preserve all existing content unless the new source directly contradicts it
- Fill in sections marked "> *To be populated*" when the source has relevant data
- Add information under the correct section; do not create new top-level sections unless essential
- If a source contradicts existing data, use the newer source and note the conflict
- Keep the document structured, business-readable, and concise — no filler
- Do NOT include source file names or meta-commentary in the overview content itself
- Return ONLY the full updated ananas-overview.md content, nothing else
- Update the comment at the top: <!-- Last updated: YYYY-MM-DD | Sources processed: N -->"""


def call_claude(file_path: Path, extracted: dict) -> str:
    import anthropic

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY not set in environment")

    client = anthropic.Anthropic(api_key=api_key)
    current_overview = OVERVIEW.read_text(encoding="utf-8")

    # count existing processed sources
    sources_text = SOURCES.read_text(encoding="utf-8")
    import re
    processed_count = len(re.findall(r"^\| ", sources_text, re.MULTILINE))

    user_content = [
        {
            "type": "text",
            "text": (
                f"Current ananas-overview.md:\n\n```markdown\n{current_overview}\n```\n\n"
                f"New source file: `{file_path.name}`\n\n"
                "Extract all company-relevant information from this source and return the fully updated overview."
            ),
        }
    ]

    if extracted["type"] == "image":
        user_content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": extracted["media_type"],
                "data": extracted["content"],
            },
        })
    else:
        user_content.append({
            "type": "text",
            "text": f"Source content:\n\n{extracted['content'][:40000]}",  # cap at 40k chars
        })

    log.info(f"Calling Claude for: {file_path.name}")
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_content}],
    )

    return response.content[0].text


# ── file processing ────────────────────────────────────────────────────────────

def mark_processed(file_path: Path) -> None:
    PROCESSED_DIR.mkdir(exist_ok=True)
    dest = PROCESSED_DIR / file_path.name
    # avoid name collision
    if dest.exists():
        stem = file_path.stem
        suffix = file_path.suffix
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        dest = PROCESSED_DIR / f"{stem}_{ts}{suffix}"
    file_path.rename(dest)
    log.info(f"Moved to .processed/: {dest.name}")


def update_sources_log(file_path: Path, sections_hint: str = "") -> None:
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    entry = f"| {file_path.name} | {today} | {sections_hint or 'overview updated'} |\n"

    text = SOURCES.read_text(encoding="utf-8")
    if "## Processed\n\n*None yet.*" in text:
        # first entry — replace placeholder and add table header
        header = "## Processed\n\n| File | Processed At | Notes |\n|---|---|---|\n"
        text = text.replace("## Processed\n\n*None yet.*", header + entry)
    else:
        # find the table and append
        insert_after = "| File | Processed At | Notes |\n|---|---|---|\n"
        text = text.replace(insert_after, insert_after + entry)

    SOURCES.write_text(text, encoding="utf-8")


def process_file(file_path: Path) -> None:
    log.info(f"Processing: {file_path.name}")
    try:
        extracted = extract_content(file_path)
        updated_overview = call_claude(file_path, extracted)
        OVERVIEW.write_text(updated_overview, encoding="utf-8")
        log.info(f"ananas-overview.md updated.")
        update_sources_log(file_path)
        mark_processed(file_path)
        log.info(f"Done: {file_path.name}")
    except Exception as e:
        log.error(f"Failed to process {file_path.name}: {e}")
        raise


def process_pending() -> int:
    """Process all files currently in raw/ (excluding hidden files and subdirs)."""
    files = [
        f for f in RAW_DIR.iterdir()
        if f.is_file() and not f.name.startswith(".")
    ]
    if not files:
        log.info("No pending files in raw/.")
        return 0
    for f in sorted(files):
        process_file(f)
    return len(files)


# ── watchdog handler ───────────────────────────────────────────────────────────

class RawFolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        path = Path(event.src_path)
        if path.name.startswith("."):
            return
        # brief delay to ensure the file is fully written before reading
        time.sleep(1)
        if path.exists():
            process_file(path)

    def on_moved(self, event):
        """Handle files moved/copied into the folder."""
        if event.is_directory:
            return
        path = Path(event.dest_path)
        if path.parent == RAW_DIR and not path.name.startswith("."):
            time.sleep(1)
            if path.exists():
                process_file(path)


# ── entrypoint ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Watch context/ananas/raw/ for new files.")
    parser.add_argument(
        "--once",
        action="store_true",
        help="Process all pending files and exit (no continuous watch).",
    )
    args = parser.parse_args()

    if args.once:
        count = process_pending()
        log.info(f"Processed {count} file(s). Exiting.")
        return

    log.info(f"Watching: {RAW_DIR}")
    log.info("Drop files into context/ananas/raw/ — they will be processed automatically.")
    log.info("Press Ctrl+C to stop.\n")

    # process anything already sitting in the folder before we start watching
    process_pending()

    observer = Observer()
    observer.schedule(RawFolderHandler(), str(RAW_DIR), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Stopping watcher.")
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
